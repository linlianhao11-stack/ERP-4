"""供应商路由"""
from __future__ import annotations

from decimal import Decimal
from datetime import datetime
from io import BytesIO
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from tortoise import transactions
from tortoise.expressions import F

from app.auth.dependencies import require_permission
from app.models import User, Supplier, PurchaseOrder, RebateLog
from app.models.supplier_balance import SupplierAccountBalance
from app.schemas.supplier import SupplierRequest, CreditRefundRequest
from app.services.operation_log_service import log_operation
from app.utils.response import paginated_response
from app.logger import get_logger

logger = get_logger("suppliers")

router = APIRouter(prefix="/api/suppliers", tags=["供应商管理"])


# ── 批量导入相关端点（放在 /{id} 路径之前，避免被拦截） ──────────────


@router.post("/import")
async def import_suppliers(
    file: UploadFile = File(...),
    user: User = Depends(require_permission("dropship")),
):
    """供应商批量导入(Excel)"""
    # 1. 校验文件后缀
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 格式的文件")

    # 2. 读取 Excel
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents), read_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="无法解析 Excel 文件，请确认格式正确")

    # 3. 列映射（按顺序）：名称, 联系人, 电话, 地址, 税号, 银行名称, 银行账号
    field_mapping = [
        "name", "contact_person", "phone", "address",
        "tax_id", "bank_name", "bank_account",
    ]

    created = 0
    skipped = 0
    errors: list[dict] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # 跳过全空行
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        # 解析行数据
        row_data: dict = {}
        for col_idx, field in enumerate(field_mapping):
            value = row[col_idx] if col_idx < len(row) else None
            if value is not None:
                row_data[field] = str(value).strip()

        # 4. 名称必填校验
        name = row_data.get("name", "")
        if not name:
            errors.append({"row": row_idx, "reason": "名称为空"})
            continue

        # 5. 去重：按名称匹配，已存在则跳过
        exists = await Supplier.filter(name=name).exists()
        if exists:
            skipped += 1
            continue

        # 6. 创建供应商
        try:
            await Supplier.create(**row_data)
            created += 1
        except Exception as e:
            errors.append({"row": row_idx, "reason": str(e)})

    wb.close()

    return {"created": created, "skipped": skipped, "errors": errors}


@router.get("/import-template")
async def import_template(
    user: User = Depends(require_permission("dropship")),
):
    """下载供应商导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "供应商导入模板"

    # 表头行
    ws.append(["名称", "联系人", "电话", "地址", "税号", "银行名称", "银行账号"])
    # 示例数据
    ws.append([
        "示例供应商", "张三", "13800138000",
        "上海市浦东新区", "91310000XXXXXXXX", "中国银行", "6217001234567890",
    ])

    # 设置列宽
    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=supplier_import_template.xlsx"
        },
    )


@router.get("")
async def list_suppliers(account_set_id: Optional[int] = None, user: User = Depends(require_permission("purchase"))):
    suppliers = await Supplier.filter(is_active=True).order_by("-created_at")

    # 如果指定账套，从 SupplierAccountBalance 读取余额
    balance_map = {}
    if account_set_id:
        balances = await SupplierAccountBalance.filter(account_set_id=account_set_id).all()
        balance_map = {b.supplier_id: b for b in balances}

    items = [{"id": s.id, "name": s.name, "contact_person": s.contact_person, "phone": s.phone,
             "tax_id": s.tax_id, "bank_account": s.bank_account, "bank_name": s.bank_name,
             "address": s.address,
             "rebate_balance": float(balance_map[s.id].rebate_balance) if s.id in balance_map else (0 if account_set_id else float(s.rebate_balance)),
             "credit_balance": float(balance_map[s.id].credit_balance) if s.id in balance_map else (0 if account_set_id else float(s.credit_balance)),
             "created_at": s.created_at.isoformat()} for s in suppliers]
    return paginated_response(items)


@router.post("")
async def create_supplier(data: SupplierRequest, user: User = Depends(require_permission("purchase"))):
    s = await Supplier.create(**data.model_dump())
    await log_operation(user, "SUPPLIER_CREATE", "SUPPLIER", s.id, f"新增供应商 {data.name}")
    return {"id": s.id, "message": "创建成功"}


@router.put("/{supplier_id}")
async def update_supplier(supplier_id: int, data: SupplierRequest, user: User = Depends(require_permission("purchase"))):
    s = await Supplier.filter(id=supplier_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="供应商不存在")
    update_data = data.model_dump(exclude_unset=True)
    if update_data:
        await Supplier.filter(id=supplier_id).update(**update_data)
    return {"message": "更新成功"}


@router.delete("/{supplier_id}")
async def delete_supplier(supplier_id: int, user: User = Depends(require_permission("purchase"))):
    s = await Supplier.filter(id=supplier_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="供应商不存在")
    # 检查是否有任何账套下的未结清余额
    from tortoise.queryset import Q
    has_balance = await SupplierAccountBalance.filter(
        Q(rebate_balance__gt=0) | Q(credit_balance__gt=0),
        supplier_id=supplier_id
    ).exists()
    if has_balance:
        raise HTTPException(status_code=400, detail="供应商有未结清返利或在账资金，无法删除")
    pending_count = await PurchaseOrder.filter(
        supplier_id=supplier_id, status__in=["pending_review", "pending", "paid", "partial"]
    ).count()
    if pending_count > 0:
        raise HTTPException(status_code=400, detail=f"该供应商有 {pending_count} 个未完成的采购单，无法删除")
    s.is_active = False
    await s.save()
    return {"message": "删除成功"}


@router.get("/{supplier_id}/transactions")
async def get_supplier_transactions(
    supplier_id: int,
    month: Optional[str] = None,
    account_set_id: Optional[int] = None,
    user: User = Depends(require_permission("purchase"))
):
    supplier = await Supplier.filter(id=supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="供应商不存在")

    # 按账套读取余额
    rebate_balance = 0
    credit_balance = 0
    if account_set_id:
        bal = await SupplierAccountBalance.filter(
            supplier_id=supplier_id, account_set_id=account_set_id
        ).first()
        if bal:
            rebate_balance = float(bal.rebate_balance)
            credit_balance = float(bal.credit_balance)
    else:
        rebate_balance = float(supplier.rebate_balance)
        credit_balance = float(supplier.credit_balance)

    # 采购记录查询
    po_query = PurchaseOrder.filter(supplier_id=supplier_id)
    if month:
        try:
            year, mon = month.split("-")
            start = datetime(int(year), int(mon), 1)
            if int(mon) == 12:
                end = datetime(int(year) + 1, 1, 1)
            else:
                end = datetime(int(year), int(mon) + 1, 1)
            po_query = po_query.filter(created_at__gte=start, created_at__lt=end)
        except (ValueError, TypeError) as e:
            logger.warning(f"供应商交易记录月份参数解析失败: month={month}, {e}")

    orders = await po_query.order_by("-created_at").limit(1000).select_related("creator")

    # 统计
    from tortoise.functions import Count, Sum, Coalesce
    stats_result = await PurchaseOrder.filter(supplier_id=supplier_id).annotate(
        total_count=Count("id"),
        total_amount=Coalesce(Sum("total_amount"), 0),
    ).values("total_count", "total_amount")
    total_count = stats_result[0]["total_count"] if stats_result else 0
    total_amount = float(stats_result[0]["total_amount"]) if stats_result else 0

    completed_count = await PurchaseOrder.filter(supplier_id=supplier_id, status="completed").count()

    returned_result = await PurchaseOrder.filter(supplier_id=supplier_id, status="returned").annotate(
        cnt=Count("id"),
        amt=Coalesce(Sum("return_amount"), 0),
    ).values("cnt", "amt")
    returned_count = returned_result[0]["cnt"] if returned_result else 0
    returned_amount = float(returned_result[0]["amt"]) if returned_result else 0

    # 在账资金流水
    credit_log_query = RebateLog.filter(
        target_type="supplier", target_id=supplier_id,
        type__in=["credit_charge", "credit_use", "credit_refund"]
    )
    if account_set_id:
        credit_log_query = credit_log_query.filter(account_set_id=account_set_id)
    credit_logs = await credit_log_query.order_by("-created_at").select_related("creator")

    from tortoise import connections
    conn = connections.get("default")
    month_rows = await conn.execute_query_dict(
        "SELECT DISTINCT TO_CHAR(created_at, 'YYYY-MM') as month FROM purchase_orders WHERE supplier_id = $1 ORDER BY month DESC",
        [supplier_id]
    )
    available_months = [r["month"] for r in month_rows]

    return {
        "supplier": {
            "id": supplier.id, "name": supplier.name,
            "contact_person": supplier.contact_person,
            "phone": supplier.phone,
            "rebate_balance": rebate_balance,
            "credit_balance": credit_balance,
        },
        "stats": {
            "total_count": total_count,
            "total_amount": round(total_amount, 2),
            "completed_count": completed_count,
            "returned_count": returned_count,
            "returned_amount": round(returned_amount, 2),
        },
        "orders": [{
            "id": o.id, "po_no": o.po_no, "status": o.status,
            "total_amount": float(o.total_amount),
            "return_amount": float(o.return_amount) if o.return_amount else 0,
            "creator_name": o.creator.display_name if o.creator else None,
            "created_at": o.created_at.isoformat(),
        } for o in orders],
        "credit_logs": [{
            "id": log.id, "type": log.type,
            "amount": float(log.amount),
            "balance_after": float(log.balance_after),
            "remark": log.remark,
            "creator_name": log.creator.display_name if log.creator else None,
            "created_at": log.created_at.isoformat(),
        } for log in credit_logs],
        "available_months": available_months,
    }


@router.post("/{supplier_id}/credit-refund")
async def refund_supplier_credit(
    supplier_id: int,
    data: CreditRefundRequest,
    user: User = Depends(require_permission("purchase"))
):
    amount = Decimal(str(data.amount))
    account_set_id = getattr(data, 'account_set_id', None)

    async with transactions.in_transaction():
        supplier = await Supplier.filter(id=supplier_id).first()
        if not supplier:
            raise HTTPException(status_code=404, detail="供应商不存在")

        if account_set_id:
            bal = await SupplierAccountBalance.filter(
                supplier_id=supplier_id, account_set_id=account_set_id
            ).select_for_update().first()
            if not bal or amount > bal.credit_balance:
                available = float(bal.credit_balance) if bal else 0
                raise HTTPException(status_code=400, detail=f"退款金额超过在账资金余额（可用: ¥{available:.2f}）")
            await SupplierAccountBalance.filter(id=bal.id).update(
                credit_balance=F('credit_balance') - amount
            )
            await bal.refresh_from_db()
            balance_after = bal.credit_balance
        else:
            # 兼容旧逻辑
            supplier = await Supplier.filter(id=supplier_id).select_for_update().first()
            if amount > supplier.credit_balance:
                raise HTTPException(status_code=400, detail=f"退款金额超过在账资金余额（可用: ¥{float(supplier.credit_balance):.2f}）")
            await Supplier.filter(id=supplier_id).update(credit_balance=F('credit_balance') - amount)
            await supplier.refresh_from_db()
            balance_after = supplier.credit_balance

        await RebateLog.create(
            target_type="supplier", target_id=supplier_id,
            type="credit_refund", amount=-amount,
            balance_after=balance_after,
            account_set_id=account_set_id,
            remark=data.remark or f"在账资金退款 ¥{float(amount):.2f}",
            creator=user
        )

        await log_operation(user, "CREDIT_REFUND", "SUPPLIER", supplier_id,
            f"供应商 {supplier.name} 在账资金退款 ¥{float(amount):.2f}，余额 ¥{float(balance_after):.2f}")

    return {"message": "退款成功", "credit_balance": float(balance_after)}
