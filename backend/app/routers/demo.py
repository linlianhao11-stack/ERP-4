"""样机管理路由"""
from __future__ import annotations

import io
from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from urllib.parse import quote

from app.auth.dependencies import require_permission
from app.models import User, DemoUnit, DemoLoan, DemoDisposal, Customer, Employee, Location
from app.schemas.demo import (
    DemoUnitCreate, DemoUnitUpdate,
    DemoLoanCreate, DemoLoanReturn,
    DemoSellRequest, DemoConvertRequest, DemoScrapRequest, DemoLossRequest,
)
from app.services.demo_service import (
    create_demo_unit, create_demo_loan,
    approve_demo_loan, reject_demo_loan,
    lend_demo_unit, return_demo_unit,
    sell_demo_unit, convert_demo_unit, scrap_demo_unit, report_loss_demo_unit,
    delete_demo_unit,
    get_demo_stats,
)
from app.services.operation_log_service import log_operation
from app.utils.response import paginated_response
from app.utils.time import to_naive
from app.logger import get_logger

logger = get_logger("demo_router")

router = APIRouter(prefix="/api/demo", tags=["样机管理"])

# ── 标签映射 ──

STATUS_LABELS = {
    "in_stock": "在库", "lent_out": "借出中", "repairing": "维修中",
    "sold": "已转销售", "scrapped": "已报废", "lost": "已丢失", "converted": "已转良品",
}
CONDITION_LABELS = {"new": "全新", "good": "良好", "fair": "一般", "poor": "较差"}
LOAN_TYPE_LABELS = {"customer_trial": "客户试用", "salesperson": "业务员携带", "exhibition": "展会"}
LOAN_STATUS_LABELS = {
    "pending_approval": "待审批", "approved": "已审批", "rejected": "已拒绝",
    "lent_out": "借出中", "returned": "已归还", "closed": "已关闭",
}


# ── 序列化 ──

async def _serialize_unit(u: DemoUnit) -> dict:
    """样机序列化，需要预加载 product / warehouse / sn_code"""
    # 解析仓位（nullable FK 不能用 select_related，单独查询）
    location_code = None
    location_name = None
    if u.location_id:
        loc = await Location.filter(id=u.location_id).first()
        if loc:
            location_code = loc.code
            location_name = loc.name

    # 获取持有人名称
    holder_name = None
    if u.current_holder_id:
        if u.current_holder_type == "customer":
            c = await Customer.filter(id=u.current_holder_id).first()
            holder_name = c.name if c else None
        elif u.current_holder_type == "employee":
            e = await Employee.filter(id=u.current_holder_id).first()
            holder_name = e.name if e else None

    # 查询当前活跃借出记录
    current_loan_id = None
    current_loan_date = None
    days_borrowed = 0
    is_overdue = False
    if u.status == "lent_out":
        active_loan = await DemoLoan.filter(demo_unit_id=u.id, status="lent_out").first()
        if active_loan:
            current_loan_id = active_loan.id
            current_loan_date = active_loan.loan_date.isoformat() if active_loan.loan_date else None
            if active_loan.loan_date:
                days_borrowed = (date.today() - active_loan.loan_date).days
            if active_loan.expected_return_date and active_loan.expected_return_date < date.today():
                is_overdue = True

    return {
        "id": u.id,
        "code": u.code,
        "product_id": u.product_id,
        "product_name": u.product.name if u.product else None,
        "product_sku": u.product.sku if u.product else None,
        "sn_code_id": u.sn_code_id,
        "sn_code": u.sn_code.sn_code if u.sn_code else None,
        "warehouse_id": u.warehouse_id,
        "warehouse_name": u.warehouse.name if u.warehouse else None,
        "location_id": u.location_id,
        "location_code": location_code,
        "location_name": location_name,
        "status": u.status,
        "status_label": STATUS_LABELS.get(u.status, u.status),
        "condition": u.condition,
        "condition_label": CONDITION_LABELS.get(u.condition, u.condition),
        "cost_price": float(u.cost_price),
        "current_holder_type": u.current_holder_type,
        "current_holder_id": u.current_holder_id,
        "holder_name": holder_name,
        "current_loan_id": current_loan_id,
        "current_loan_date": current_loan_date,
        "days_borrowed": days_borrowed,
        "is_overdue": is_overdue,
        "total_loan_count": u.total_loan_count,
        "total_loan_days": u.total_loan_days,
        "notes": u.notes,
        "created_at": u.created_at.isoformat() if u.created_at else None,
    }


async def _serialize_loan(loan: DemoLoan) -> dict:
    """借出记录序列化"""
    # 获取借用人名称
    borrower_name = None
    if loan.borrower_type == "customer":
        c = await Customer.filter(id=loan.borrower_id).first()
        borrower_name = c.name if c else None
    elif loan.borrower_type == "employee":
        e = await Employee.filter(id=loan.borrower_id).first()
        borrower_name = e.name if e else None

    # 经手人
    handler_name = None
    if loan.handler_id:
        h = await Employee.filter(id=loan.handler_id).first()
        handler_name = h.name if h else None

    # 逾期判定
    is_overdue = (
        loan.status == "lent_out"
        and loan.expected_return_date
        and loan.expected_return_date < date.today()
    )

    # 产品名称（通过 demo_unit 关联）
    product_name = None
    unit = loan.demo_unit if hasattr(loan, "demo_unit") and loan.demo_unit else None
    if unit:
        try:
            await unit.fetch_related("product")
            product_name = unit.product.name if unit.product else None
        except Exception:
            pass

    return {
        "id": loan.id,
        "loan_no": loan.loan_no,
        "demo_unit_id": loan.demo_unit_id,
        "demo_unit_code": unit.code if unit else None,
        "product_name": product_name,
        "loan_type": loan.loan_type,
        "loan_type_label": LOAN_TYPE_LABELS.get(loan.loan_type, loan.loan_type),
        "borrower_type": loan.borrower_type,
        "borrower_id": loan.borrower_id,
        "borrower_name": borrower_name,
        "handler_id": loan.handler_id,
        "handler_name": handler_name,
        "status": loan.status,
        "status_label": LOAN_STATUS_LABELS.get(loan.status, loan.status),
        "loan_date": loan.loan_date.isoformat() if loan.loan_date else None,
        "expected_return_date": loan.expected_return_date.isoformat() if loan.expected_return_date else None,
        "actual_return_date": loan.actual_return_date.isoformat() if loan.actual_return_date else None,
        "condition_on_loan": loan.condition_on_loan,
        "condition_on_return": loan.condition_on_return,
        "return_notes": loan.return_notes,
        "purpose": loan.purpose,
        "is_overdue": is_overdue,
        "created_at": loan.created_at.isoformat() if loan.created_at else None,
    }


# ── 统计 ──

@router.get("/stats")
async def stats(
    warehouse_id: Optional[int] = None,
    user: User = Depends(require_permission("stock_view")),
):
    return await get_demo_stats(warehouse_id)


# ── 样机 CRUD ──

@router.get("/units/export")
async def export_units(user: User = Depends(require_permission("stock_view"))):
    """导出样机列表为 Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        units = await DemoUnit.all().select_related("product", "warehouse", "sn_code").order_by("code")

        wb = Workbook()
        ws = wb.active
        ws.title = "样机列表"

        headers = [
            "样机编号", "商品名称", "SKU", "SN码", "所属仓库", "仓位",
            "状态", "成色", "成本价", "当前持有人",
            "累计借出次数", "累计借出天数", "备注", "入库时间",
        ]

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # 列宽
        col_widths = [14, 24, 16, 22, 14, 10, 10, 8, 12, 14, 12, 12, 24, 18]
        from openpyxl.utils import get_column_letter
        for idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = w

        for row_idx, u in enumerate(units, 2):
            # 持有人名称
            holder_name = ""
            if u.current_holder_id:
                if u.current_holder_type == "customer":
                    c = await Customer.filter(id=u.current_holder_id).first()
                    holder_name = c.name if c else ""
                elif u.current_holder_type == "employee":
                    e = await Employee.filter(id=u.current_holder_id).first()
                    holder_name = e.name if e else ""

            created = ""
            if u.created_at:
                dt = to_naive(u.created_at)
                created = dt.strftime("%Y-%m-%d %H:%M") if dt else ""

            location_display = ""
            if u.location_id:
                loc = await Location.filter(id=u.location_id).first()
                if loc:
                    location_display = f"{loc.code} - {loc.name}" if loc.name else loc.code

            row_data = [
                u.code,
                u.product.name if u.product else "",
                u.product.sku if u.product else "",
                u.sn_code.sn_code if u.sn_code else "",
                u.warehouse.name if u.warehouse else "",
                location_display,
                STATUS_LABELS.get(u.status, u.status),
                CONDITION_LABELS.get(u.condition, u.condition),
                float(u.cost_price),
                holder_name,
                u.total_loan_count,
                u.total_loan_days,
                u.notes or "",
                created,
            ]
            for col_idx, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        from datetime import datetime
        filename = f"样机列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        await log_operation(user, "DEMO_EXPORT", "DEMO_UNIT", None, "导出样机列表 Excel")
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
        )
    except Exception as e:
        logger.error("样机导出失败", exc_info=e)
        raise HTTPException(status_code=500, detail="导出失败，请重试")


@router.get("/units")
async def list_units(
    status: Optional[str] = None,
    search: Optional[str] = None,
    warehouse_id: Optional[int] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    user: User = Depends(require_permission("stock_view")),
):
    """样机列表（分页 + 筛选）"""
    query = DemoUnit.all()
    if status:
        query = query.filter(status=status)
    if warehouse_id:
        query = query.filter(warehouse_id=warehouse_id)
    if search:
        from tortoise.queryset import Q
        q = Q(code__icontains=search) | Q(product__name__icontains=search) | Q(product__sku__icontains=search)
        query = DemoUnit.all().filter(q)
        if status:
            query = query.filter(status=status)
        if warehouse_id:
            query = query.filter(warehouse_id=warehouse_id)

    total = await query.count()
    units = await query.order_by("-created_at").offset((page - 1) * page_size).limit(page_size).select_related(
        "product", "warehouse", "sn_code",
    )
    items = [await _serialize_unit(u) for u in units]
    return paginated_response(items, total)


@router.post("/units")
async def create_unit(
    data: DemoUnitCreate,
    user: User = Depends(require_permission("stock_edit")),
):
    unit = await create_demo_unit(data, user)
    await log_operation(user, "DEMO_CREATE", "DEMO_UNIT", unit.id, f"新建样机 {unit.code}")
    # 重新加载关联
    unit = await DemoUnit.filter(id=unit.id).select_related("product", "warehouse", "sn_code").first()
    return await _serialize_unit(unit)


@router.get("/units/{unit_id}")
async def get_unit(
    unit_id: int,
    user: User = Depends(require_permission("stock_view")),
):
    unit = await DemoUnit.filter(id=unit_id).select_related("product", "warehouse", "sn_code").first()
    if not unit:
        raise HTTPException(status_code=404, detail="样机不存在")

    result = await _serialize_unit(unit)

    # 借出历史
    loans = await DemoLoan.filter(demo_unit_id=unit_id).select_related("demo_unit").order_by("-created_at").limit(50)
    result["loans"] = [await _serialize_loan(l) for l in loans]

    # 处置记录
    disposals = await DemoDisposal.filter(demo_unit_id=unit_id).order_by("-created_at")
    result["disposals"] = [{
        "id": d.id,
        "disposal_type": d.disposal_type,
        "amount": float(d.amount) if d.amount else None,
        "refurbish_cost": float(d.refurbish_cost) if d.refurbish_cost else None,
        "reason": d.reason,
        "created_at": d.created_at.isoformat() if d.created_at else None,
    } for d in disposals]

    return result


@router.put("/units/{unit_id}")
async def update_unit(
    unit_id: int,
    data: DemoUnitUpdate,
    user: User = Depends(require_permission("stock_edit")),
):
    unit = await DemoUnit.filter(id=unit_id).first()
    if not unit:
        raise HTTPException(status_code=404, detail="样机不存在")

    if data.condition is not None:
        unit.condition = data.condition
    if data.notes is not None:
        unit.notes = data.notes
    await unit.save()
    await log_operation(user, "DEMO_UPDATE", "DEMO_UNIT", unit.id, f"更新样机 {unit.code}")

    unit = await DemoUnit.filter(id=unit_id).select_related("product", "warehouse", "sn_code").first()
    return await _serialize_unit(unit)


@router.delete("/units/{unit_id}")
async def delete_unit(
    unit_id: int,
    user: User = Depends(require_permission("stock_edit")),
):
    await delete_demo_unit(unit_id, user)
    await log_operation(user, "DEMO_DELETE", "DEMO_UNIT", unit_id, f"删除样机 ID={unit_id}")
    return {"ok": True}


# ── 借还 ──

@router.get("/loans")
async def list_loans(
    status: Optional[str] = None,
    search: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    user: User = Depends(require_permission("stock_view")),
):
    """借出记录列表"""
    is_overdue_filter = status == "overdue"

    query = DemoLoan.all()
    if status and not is_overdue_filter:
        query = query.filter(status=status)
    if is_overdue_filter:
        query = query.filter(status="lent_out", expected_return_date__lt=date.today())
    if start_date:
        query = query.filter(created_at__gte=start_date)
    if end_date:
        from datetime import datetime, time
        end_dt = datetime.combine(end_date, time.max)
        query = query.filter(created_at__lte=end_dt)
    if search:
        from tortoise.queryset import Q
        q = Q(loan_no__icontains=search) | Q(demo_unit__code__icontains=search)
        query = query.filter(q)

    total = await query.count()
    loans = await query.offset((page - 1) * page_size).limit(page_size).select_related("demo_unit").order_by("-created_at")
    items = [await _serialize_loan(l) for l in loans]
    return paginated_response(items, total)


@router.post("/loans")
async def create_loan(
    data: DemoLoanCreate,
    user: User = Depends(require_permission("stock_edit")),
):
    loan = await create_demo_loan(data, user)
    await log_operation(user, "DEMO_LOAN_CREATE", "DEMO_UNIT", data.demo_unit_id, f"创建借出申请 {loan.loan_no}")
    loan = await DemoLoan.filter(id=loan.id).select_related("demo_unit").first()
    return await _serialize_loan(loan)


@router.get("/loans/{loan_id}")
async def get_loan(
    loan_id: int,
    user: User = Depends(require_permission("stock_view")),
):
    loan = await DemoLoan.filter(id=loan_id).select_related("demo_unit").first()
    if not loan:
        raise HTTPException(status_code=404, detail="借出记录不存在")
    return await _serialize_loan(loan)


@router.post("/loans/{loan_id}/approve")
async def approve_loan(
    loan_id: int,
    user: User = Depends(require_permission("stock_edit")),
):
    loan = await approve_demo_loan(loan_id, user)
    await log_operation(user, "DEMO_LOAN_APPROVE", "DEMO_UNIT", loan.demo_unit_id, f"审批通过借出申请 {loan.loan_no}")
    loan = await DemoLoan.filter(id=loan.id).select_related("demo_unit").first()
    return await _serialize_loan(loan)


@router.post("/loans/{loan_id}/reject")
async def reject_loan(
    loan_id: int,
    user: User = Depends(require_permission("stock_edit")),
):
    loan = await reject_demo_loan(loan_id, user)
    await log_operation(user, "DEMO_LOAN_REJECT", "DEMO_UNIT", loan.demo_unit_id, f"拒绝借出申请 {loan.loan_no}")
    loan = await DemoLoan.filter(id=loan.id).select_related("demo_unit").first()
    return await _serialize_loan(loan)


@router.post("/loans/{loan_id}/lend")
async def lend_unit(
    loan_id: int,
    user: User = Depends(require_permission("stock_edit")),
):
    loan = await lend_demo_unit(loan_id, user)
    await log_operation(user, "DEMO_LEND", "DEMO_UNIT", loan.demo_unit_id, f"样机借出 {loan.loan_no}")
    loan = await DemoLoan.filter(id=loan.id).select_related("demo_unit").first()
    return await _serialize_loan(loan)


@router.post("/loans/{loan_id}/return")
async def return_unit(
    loan_id: int,
    data: DemoLoanReturn,
    user: User = Depends(require_permission("stock_edit")),
):
    loan = await return_demo_unit(loan_id, data, user)
    await log_operation(user, "DEMO_RETURN", "DEMO_UNIT", loan.demo_unit_id, f"样机归还 {loan.loan_no}")
    loan = await DemoLoan.filter(id=loan.id).select_related("demo_unit").first()
    return await _serialize_loan(loan)


# ── 处置 ──

@router.post("/units/{unit_id}/sell")
async def sell_unit(
    unit_id: int,
    data: DemoSellRequest,
    user: User = Depends(require_permission("stock_edit")),
):
    disposal = await sell_demo_unit(unit_id, data, user)
    await log_operation(user, "DEMO_SELL", "DEMO_UNIT", unit_id, f"样机转销售 ID={unit_id}")
    return {"message": "转销售成功", "disposal_id": disposal.id, "order_id": disposal.order_id}


@router.post("/units/{unit_id}/convert")
async def convert_unit(
    unit_id: int,
    data: DemoConvertRequest,
    user: User = Depends(require_permission("stock_edit")),
):
    disposal = await convert_demo_unit(unit_id, data, user)
    await log_operation(user, "DEMO_CONVERT", "DEMO_UNIT", unit_id, f"样机转良品 ID={unit_id}")
    return {"message": "转良品成功", "disposal_id": disposal.id}


@router.post("/units/{unit_id}/scrap")
async def scrap_unit(
    unit_id: int,
    data: DemoScrapRequest,
    user: User = Depends(require_permission("stock_edit")),
):
    disposal = await scrap_demo_unit(unit_id, data, user)
    await log_operation(user, "DEMO_SCRAP", "DEMO_UNIT", unit_id, f"样机报废 ID={unit_id}")
    return {"message": "报废成功", "disposal_id": disposal.id}


@router.post("/units/{unit_id}/report-loss")
async def report_loss(
    unit_id: int,
    data: DemoLossRequest,
    user: User = Depends(require_permission("stock_edit")),
):
    disposal = await report_loss_demo_unit(unit_id, data, user)
    await log_operation(user, "DEMO_LOSS", "DEMO_UNIT", unit_id, f"样机丢失报告 ID={unit_id}")
    return {"message": "丢失报告已提交", "disposal_id": disposal.id}
