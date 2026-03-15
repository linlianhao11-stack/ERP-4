"""AI 聊天路由 — 聊天 / 导出 / 反馈 / 状态"""
from __future__ import annotations
import asyncio
import io
import json
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from app.auth.dependencies import get_current_user, require_permission
from app.models import User, SystemSetting
from app.schemas.ai import ChatRequest, FeedbackRequest, ExportRequest
from app.services.ai_chat_service import process_chat, check_ai_available, get_preset_queries
from app.ai.rate_limiter import user_limiter, global_limiter
from app.config import DATABASE_URL, AI_DB_PASSWORD
from app.logger import get_logger

logger = get_logger("ai.chat")
router = APIRouter(prefix="/api/ai", tags=["AI助手"])


def _build_ai_dsn() -> str:
    """构建 AI 只读用户的数据库连接字符串"""
    # 从主 DSN 解析 host/port/dbname
    # DATABASE_URL 格式: postgres://user:pass@host:port/dbname
    import re
    from urllib.parse import quote_plus
    match = re.match(r'postgres(?:ql)?://[^@]+@([^/]+)/([\w-]+)', DATABASE_URL)
    if not match:
        raise RuntimeError("无法解析 DATABASE_URL")
    host_port = match.group(1)
    dbname = match.group(2)
    if not AI_DB_PASSWORD:
        raise RuntimeError("AI_DB_PASSWORD 环境变量未设置，无法连接 AI 只读数据库")
    return f"postgresql://erp_ai_readonly:{quote_plus(AI_DB_PASSWORD)}@{host_port}/{dbname}"


@router.get("/status")
async def ai_status(user: User = Depends(get_current_user)):
    """检查 AI 助手可用性（含预设快捷问题，所有登录用户可访问）"""
    available, reason = await check_ai_available()
    if available and not user.has_permission("ai_chat"):
        available = False
        reason = "你没有 AI 助手使用权限"
    perms = None if user.role == "admin" else (user.permissions or [])
    preset = await get_preset_queries(user_permissions=perms) if available else []
    return {"available": available, "reason": reason, "preset_queries": preset}


@router.post("/chat")
async def ai_chat(body: ChatRequest, user: User = Depends(require_permission("ai_chat"))):
    """AI 聊天接口 — SSE 流式响应"""
    import unicodedata

    # 输入验证 — history 结构 + 大小
    if len(body.history) > 50:
        body.history = body.history[-50:]
    valid_roles = {"user", "assistant"}
    total_size = 0
    for h in body.history:
        if not isinstance(h, dict) or h.get("role") not in valid_roles:
            raise HTTPException(status_code=400, detail="无效的对话历史格式")
        content = h.get("content", "")
        if len(content) > 5000:
            h["content"] = content[:5000]
        total_size += len(h.get("content", ""))
    if total_size > 200000:  # 200KB 总大小限制
        raise HTTPException(status_code=400, detail="对话历史过长，请清空会话后重试")

    # 限流检查
    user_key = f"user:{user.id}"
    if not user_limiter.allow(user_key):
        raise HTTPException(status_code=429, detail="请求过于频繁，请稍后再试")
    if not global_limiter.allow("global"):
        raise HTTPException(status_code=429, detail="系统繁忙，请稍后再试")

    # 清洗输入 — 长度限制 + Unicode 控制字符过滤
    clean_msg = body.message.strip()
    if len(clean_msg) > 2000:
        clean_msg = clean_msg[:2000]
    clean_msg = "".join(
        c for c in clean_msg
        if c == "\n" or (not unicodedata.category(c).startswith("C"))
    )
    if not clean_msg:
        raise HTTPException(status_code=400, detail="请输入查询内容")

    dsn = _build_ai_dsn()

    async def event_stream():
        try:
            async for event in process_chat(
                message=clean_msg,
                history=body.history,
                user_id=user.id,
                db_dsn=dsn,
                user_permissions=None if user.role == "admin" else (user.permissions or []),
                is_preset=body.is_preset,
            ):
                evt_type = event.get("event", "progress")
                evt_data = json.dumps(event.get("data", {}), ensure_ascii=False)
                yield f"event: {evt_type}\ndata: {evt_data}\n\n"
        except asyncio.CancelledError:
            pass

        logger.info(f"AI 查询: user={user.username}, question={clean_msg[:50]}")

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@router.post("/chat/export")
async def ai_export(body: ExportRequest, user: User = Depends(require_permission("ai_chat"))):
    """导出查询结果为 Excel"""
    import openpyxl

    table_data = body.table_data
    if not table_data or "columns" not in table_data or "rows" not in table_data:
        raise HTTPException(status_code=400, detail="无效的表格数据")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = body.title[:31]  # Excel sheet 名最多 31 字符

    # 写表头
    for col, name in enumerate(table_data["columns"], 1):
        ws.cell(row=1, column=col, value=name)

    # 写数据行
    for row_idx, row in enumerate(table_data["rows"], 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # 输出为字节流
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{body.title}.xlsx"
    from urllib.parse import quote
    encoded_name = quote(filename)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"export.xlsx\"; filename*=UTF-8''{encoded_name}"},
    )


@router.post("/chat/feedback")
async def ai_feedback(body: FeedbackRequest, user: User = Depends(require_permission("ai_chat"))):
    """接收用户反馈"""
    logger.info(
        f"AI 反馈: user={user.username}, message_id={body.message_id}, "
        f"feedback={body.feedback}, reason={body.negative_reason}"
    )

    # 正面反馈 + 保存为示例（标记为待审批，管理员确认后才生效）
    if body.feedback == "positive" and body.save_as_example and body.original_question and body.sql:
        try:
            from tortoise.transactions import in_transaction
            async with in_transaction():
                setting = await SystemSetting.filter(key="ai.few_shots").select_for_update().first()
                shots = []
                if setting and setting.value:
                    shots = json.loads(setting.value)

                # 去重：相同问题不重复添加
                if any(s.get("question") == body.original_question for s in shots):
                    logger.info(f"few-shot 已存在，跳过: {body.original_question[:50]}")
                else:
                    new_shot = {
                        "question": body.original_question,
                        "sql": body.sql,
                        "source": "feedback",
                        "submitted_by": user.username,
                        "approved": user.role == "admin",  # admin 直接生效，普通用户待审批
                    }
                    shots.append(new_shot)

                    # 限制 feedback 来源的示例不超过 50 条，超限移除最早的
                    MAX_FEEDBACK_SHOTS = 50
                    feedback_count = sum(1 for s in shots if s.get("source") == "feedback")
                    if feedback_count > MAX_FEEDBACK_SHOTS:
                        for idx, s in enumerate(shots):
                            if s.get("source") == "feedback":
                                shots.pop(idx)
                                break

                    store_value = json.dumps(shots, ensure_ascii=False)
                    if setting:
                        setting.value = store_value
                        await setting.save()
                    else:
                        await SystemSetting.create(key="ai.few_shots", value=store_value)

                    logger.info(f"新增 few-shot 示例 (approved={new_shot['approved']}): {body.original_question[:50]}")
        except Exception as e:
            logger.error(f"保存 few-shot 失败: {e}")

    return {"ok": True}
