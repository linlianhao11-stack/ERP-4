"""财务报表 Excel + PDF 导出"""
from __future__ import annotations

import io
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.units import cm, mm
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from app.utils.pdf_print import _ensure_font, FONT_NAME, _FONT_REGISTERED, _fmt
from app.logger import get_logger

logger = get_logger("report_export")

# ── Excel 样式常量（复用 ledger_export 模式）──
_HEADER_FONT = Font(bold=True, size=10)
_TITLE_FONT = Font(bold=True, size=14)
_SUB_TITLE_FONT = Font(bold=True, size=11)
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")
_HEADER_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _set_header_row(ws, row_num, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER


# ═══════════════════════════════════════
# 资产负债表 Excel
# ═══════════════════════════════════════
def export_balance_sheet_excel(data: dict) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "资产负债表"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.cell(1, 1, value=f"资产负债表").font = _TITLE_FONT
    ws.cell(1, 1).alignment = _CENTER
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    ws.cell(2, 1, value=f"期间: {data['period_name']}").alignment = _CENTER

    # 资产侧
    _set_header_row(ws, 4, ["资产项目", "金额", "负债及所有者权益项目", "金额"])

    assets = data["assets"]
    liab = data["liabilities"]
    equity = data["equity"]

    # 构建左右对照行
    left_rows = [
        ("流动资产", ""),
        ("  库存现金", assets["current"]["cash"]),
        ("  银行存款", assets["current"]["bank"]),
        ("  应收账款", assets["current"]["accounts_receivable"]),
        ("  预付账款", assets["current"]["prepaid"]),
        ("  其他应收款", assets["current"]["other_receivable"]),
        ("  原材料", assets["current"]["raw_material"]),
        ("  库存商品", assets["current"]["inventory"]),
        ("  发出商品", assets["current"]["goods_in_transit"]),
        ("流动资产合计", assets["current"]["subtotal"]),
        ("", ""),
        ("非流动资产", ""),
        ("  固定资产", assets["non_current"]["fixed_assets"]),
        ("  累计折旧", assets["non_current"]["accumulated_depreciation"]),
        ("非流动资产合计", assets["non_current"]["subtotal"]),
        ("", ""),
        ("资产总计", assets["total"]),
    ]

    right_rows = [
        ("流动负债", ""),
        ("  短期借款", liab["current"]["short_term_loan"]),
        ("  应付账款", liab["current"]["accounts_payable"]),
        ("  预收账款", liab["current"]["advance_received"]),
        ("  应付职工薪酬", liab["current"]["salary_payable"]),
        ("  应交税费", liab["current"]["tax_payable"]),
        ("  其他应付款", liab["current"]["other_payable"]),
        ("流动负债合计", liab["current"]["subtotal"]),
        ("", ""),
        ("所有者权益", ""),
        ("  实收资本", equity["paid_capital"]),
        ("  盈余公积", equity["surplus_reserve"]),
        ("  本年利润", equity["current_profit"]),
        ("  未分配利润", equity["retained_earnings"]),
        ("所有者权益合计", equity["subtotal"]),
        ("", ""),
        ("负债和所有者权益总计", data["total_liabilities_equity"]),
    ]

    max_rows = max(len(left_rows), len(right_rows))
    for i in range(max_rows):
        row = 5 + i
        if i < len(left_rows):
            ws.cell(row, 1, value=left_rows[i][0])
            if left_rows[i][1]:
                ws.cell(row, 2, value=left_rows[i][1]).alignment = _RIGHT
            if left_rows[i][0] in ("流动资产合计", "非流动资产合计", "资产总计"):
                ws.cell(row, 1).font = _HEADER_FONT
                ws.cell(row, 2).font = _HEADER_FONT
        if i < len(right_rows):
            ws.cell(row, 3, value=right_rows[i][0])
            if right_rows[i][1]:
                ws.cell(row, 4, value=right_rows[i][1]).alignment = _RIGHT
            if right_rows[i][0] in ("流动负债合计", "所有者权益合计", "负债和所有者权益总计"):
                ws.cell(row, 3).font = _HEADER_FONT
                ws.cell(row, 4).font = _HEADER_FONT

    balanced = "平衡 ✓" if data["is_balanced"] else "不平衡 ✗"
    ws.cell(5 + max_rows + 1, 1, value=f"试算: {balanced}").font = Font(
        bold=True, color="008000" if data["is_balanced"] else "FF0000"
    )

    widths = [24, 16, 28, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ═══════════════════════════════════════
# 利润表 Excel
# ═══════════════════════════════════════
def export_income_statement_excel(data: dict) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "利润表"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws.cell(1, 1, value="利润表").font = _TITLE_FONT
    ws.cell(1, 1).alignment = _CENTER
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    ws.cell(2, 1, value=f"期间: {data['period_name']}").alignment = _CENTER

    _set_header_row(ws, 4, ["项目", "本期金额", "本年累计"])

    for i, r in enumerate(data["rows"], 5):
        ws.cell(i, 1, value=r["name"])
        ws.cell(i, 2, value=r["current"]).alignment = _RIGHT
        ws.cell(i, 3, value=r["ytd"]).alignment = _RIGHT
        if r["name"].startswith(("一", "二", "三", "四", "五")):
            ws.cell(i, 1).font = _HEADER_FONT
            ws.cell(i, 2).font = _HEADER_FONT
            ws.cell(i, 3).font = _HEADER_FONT

    widths = [30, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ═══════════════════════════════════════
# 现金流量表 Excel
# ═══════════════════════════════════════
def export_cash_flow_excel(data: dict) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "现金流量表"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws.cell(1, 1, value="现金流量表").font = _TITLE_FONT
    ws.cell(1, 1).alignment = _CENTER
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    ws.cell(2, 1, value=f"期间: {data['period_name']}").alignment = _CENTER

    _set_header_row(ws, 4, ["项目", "金额"])
    row = 5

    sections = [
        ("一、经营活动产生的现金流量", data["operating"]),
        ("二、投资活动产生的现金流量", data["investing"]),
        ("三、筹资活动产生的现金流量", data["financing"]),
    ]

    for title, section in sections:
        ws.cell(row, 1, value=title).font = _SUB_TITLE_FONT
        row += 1
        for item in section["items"]:
            ws.cell(row, 1, value=f"  {item['label']}")
            ws.cell(row, 2, value=item["amount"]).alignment = _RIGHT
            row += 1
        ws.cell(row, 1, value=f"  {title[:5]}净额").font = _HEADER_FONT
        ws.cell(row, 2, value=section["net"]).alignment = _RIGHT
        ws.cell(row, 2).font = _HEADER_FONT
        row += 2

    ws.cell(row, 1, value="四、现金及现金等价物净增加额").font = _SUB_TITLE_FONT
    ws.cell(row, 2, value=data["net_increase"]).alignment = _RIGHT
    ws.cell(row, 2).font = _HEADER_FONT
    row += 1
    ws.cell(row, 1, value="  加：期初现金余额")
    ws.cell(row, 2, value=data["opening_cash"]).alignment = _RIGHT
    row += 1
    ws.cell(row, 1, value="  期末现金余额").font = _HEADER_FONT
    ws.cell(row, 2, value=data["closing_cash"]).alignment = _RIGHT
    ws.cell(row, 2).font = _HEADER_FONT

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ═══════════════════════════════════════
# PDF 导出（A4 竖版）
# ═══════════════════════════════════════
def _pdf_font():
    _ensure_font()
    return FONT_NAME if _FONT_REGISTERED else "Helvetica"


def export_balance_sheet_pdf(data: dict) -> bytes:
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4
    font = _pdf_font()

    c.setFont(font, 16)
    c.drawCentredString(w / 2, h - 2 * cm, "资产负债表")
    c.setFont(font, 10)
    c.drawCentredString(w / 2, h - 2.8 * cm, f"期间: {data['period_name']}")

    y = h - 4 * cm
    c.setFont(font, 9)

    assets = data["assets"]
    liab = data["liabilities"]
    equity = data["equity"]

    left_x = 1.5 * cm
    left_val_x = 8 * cm
    right_x = 11 * cm
    right_val_x = 18 * cm

    def draw_pair(label_l, val_l, label_r, val_r, bold=False):
        nonlocal y
        if bold:
            c.setFont(font, 10)
        c.drawString(left_x, y, label_l)
        if val_l:
            c.drawRightString(left_val_x, y, _fmt(val_l))
        c.drawString(right_x, y, label_r)
        if val_r:
            c.drawRightString(right_val_x, y, _fmt(val_r))
        if bold:
            c.setFont(font, 9)
        y -= 0.5 * cm

    # 表头线
    c.line(left_x, y + 0.3 * cm, right_val_x + 1 * cm, y + 0.3 * cm)
    draw_pair("资产项目", "", "负债及所有者权益项目", "", bold=True)
    c.line(left_x, y + 0.3 * cm, right_val_x + 1 * cm, y + 0.3 * cm)

    draw_pair("流动资产", "", "流动负债", "")
    draw_pair("  库存现金", assets["current"]["cash"], "  短期借款", liab["current"]["short_term_loan"])
    draw_pair("  银行存款", assets["current"]["bank"], "  应付账款", liab["current"]["accounts_payable"])
    draw_pair("  应收账款", assets["current"]["accounts_receivable"], "  预收账款", liab["current"]["advance_received"])
    draw_pair("  预付账款", assets["current"]["prepaid"], "  应付职工薪酬", liab["current"]["salary_payable"])
    draw_pair("  其他应收款", assets["current"]["other_receivable"], "  应交税费", liab["current"]["tax_payable"])
    draw_pair("  原材料", assets["current"]["raw_material"], "  其他应付款", liab["current"]["other_payable"])
    draw_pair("  库存商品", assets["current"]["inventory"], "", "")
    draw_pair("  发出商品", assets["current"]["goods_in_transit"], "", "")
    draw_pair("流动资产合计", assets["current"]["subtotal"], "流动负债合计", liab["current"]["subtotal"], bold=True)

    draw_pair("", "", "", "")
    draw_pair("非流动资产", "", "所有者权益", "")
    draw_pair("  固定资产", assets["non_current"]["fixed_assets"], "  实收资本", equity["paid_capital"])
    draw_pair("  累计折旧", assets["non_current"]["accumulated_depreciation"], "  盈余公积", equity["surplus_reserve"])
    draw_pair("非流动资产合计", assets["non_current"]["subtotal"], "  本年利润", equity["current_profit"], bold=True)
    draw_pair("", "", "  未分配利润", equity["retained_earnings"])
    draw_pair("", "", "所有者权益合计", equity["subtotal"], bold=True)

    c.line(left_x, y + 0.3 * cm, right_val_x + 1 * cm, y + 0.3 * cm)
    draw_pair("资产总计", assets["total"], "负债和权益总计", data["total_liabilities_equity"], bold=True)
    c.line(left_x, y + 0.3 * cm, right_val_x + 1 * cm, y + 0.3 * cm)

    c.save()
    return buf.getvalue()


def export_income_statement_pdf(data: dict) -> bytes:
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4
    font = _pdf_font()

    c.setFont(font, 16)
    c.drawCentredString(w / 2, h - 2 * cm, "利润表")
    c.setFont(font, 10)
    c.drawCentredString(w / 2, h - 2.8 * cm, f"期间: {data['period_name']}")

    y = h - 4 * cm
    # 表头
    c.setFont(font, 10)
    c.line(2 * cm, y + 0.3 * cm, 19 * cm, y + 0.3 * cm)
    c.drawString(2 * cm, y, "项目")
    c.drawRightString(12 * cm, y, "本期金额")
    c.drawRightString(18 * cm, y, "本年累计")
    y -= 0.3 * cm
    c.line(2 * cm, y, 19 * cm, y)
    y -= 0.5 * cm

    c.setFont(font, 9)
    for r in data["rows"]:
        is_section = r["name"].startswith(("一", "二", "三", "四", "五"))
        if is_section:
            c.setFont(font, 10)
        c.drawString(2 * cm, y, r["name"])
        c.drawRightString(12 * cm, y, _fmt(r["current"]))
        c.drawRightString(18 * cm, y, _fmt(r["ytd"]))
        if is_section:
            c.setFont(font, 9)
        y -= 0.5 * cm

    c.line(2 * cm, y + 0.3 * cm, 19 * cm, y + 0.3 * cm)
    c.save()
    return buf.getvalue()


def export_cash_flow_pdf(data: dict) -> bytes:
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4
    font = _pdf_font()

    c.setFont(font, 16)
    c.drawCentredString(w / 2, h - 2 * cm, "现金流量表")
    c.setFont(font, 10)
    c.drawCentredString(w / 2, h - 2.8 * cm, f"期间: {data['period_name']}")

    y = h - 4 * cm
    c.setFont(font, 10)
    c.line(2 * cm, y + 0.3 * cm, 19 * cm, y + 0.3 * cm)
    c.drawString(2 * cm, y, "项目")
    c.drawRightString(18 * cm, y, "金额")
    y -= 0.3 * cm
    c.line(2 * cm, y, 19 * cm, y)
    y -= 0.6 * cm

    sections = [
        ("一、经营活动产生的现金流量", data["operating"]),
        ("二、投资活动产生的现金流量", data["investing"]),
        ("三、筹资活动产生的现金流量", data["financing"]),
    ]

    c.setFont(font, 9)
    for title, section in sections:
        c.setFont(font, 10)
        c.drawString(2 * cm, y, title)
        y -= 0.5 * cm
        c.setFont(font, 9)
        for item in section["items"]:
            c.drawString(2.5 * cm, y, item["label"])
            c.drawRightString(18 * cm, y, _fmt(item["amount"]))
            y -= 0.5 * cm
        c.setFont(font, 10)
        c.drawString(2.5 * cm, y, "小计")
        c.drawRightString(18 * cm, y, _fmt(section["net"]))
        c.setFont(font, 9)
        y -= 0.8 * cm

    c.setFont(font, 10)
    c.line(2 * cm, y + 0.3 * cm, 19 * cm, y + 0.3 * cm)
    c.drawString(2 * cm, y, "四、现金及现金等价物净增加额")
    c.drawRightString(18 * cm, y, _fmt(data["net_increase"]))
    y -= 0.5 * cm
    c.setFont(font, 9)
    c.drawString(2.5 * cm, y, f"加：期初现金余额")
    c.drawRightString(18 * cm, y, _fmt(data["opening_cash"]))
    y -= 0.5 * cm
    c.setFont(font, 10)
    c.drawString(2.5 * cm, y, f"期末现金余额")
    c.drawRightString(18 * cm, y, _fmt(data["closing_cash"]))
    c.line(2 * cm, y - 0.2 * cm, 19 * cm, y - 0.2 * cm)

    c.save()
    return buf.getvalue()
