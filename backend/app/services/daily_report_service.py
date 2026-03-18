"""每日业务日报生成与发送服务"""
from __future__ import annotations
import asyncio
import io
from datetime import datetime

from app.logger import get_logger
from app.models import SystemSetting
from app.ai.encryption import decrypt_value

logger = get_logger("daily_report")

# 配置 keys（与 router 共享）
CONFIG_KEYS = [
    "daily_report.enabled", "daily_report.send_time",
    "daily_report.recipients", "daily_report.smtp_host",
    "daily_report.smtp_port", "daily_report.smtp_user",
    "daily_report.smtp_password", "daily_report.from_email",
    "daily_report.from_name", "daily_report.last_sent_date",
]

# 8 条固定 SQL（今天 vs 昨天环比）
REPORT_QUERIES = [
    {
        "title": "销售概况（环比）",
        "sql": """
            SELECT
                '今日' AS 日期,
                COUNT(DISTINCT order_no) AS 订单数,
                ROUND(SUM(amount)::numeric, 2) AS 销售额,
                ROUND(SUM(profit)::numeric, 2) AS 毛利,
                ROUND(SUM(profit)/NULLIF(SUM(amount),0)*100, 2) AS 毛利率
            FROM vw_sales_detail WHERE order_date = CURRENT_DATE
            UNION ALL
            SELECT
                '昨日',
                COUNT(DISTINCT order_no),
                ROUND(SUM(amount)::numeric, 2),
                ROUND(SUM(profit)::numeric, 2),
                ROUND(SUM(profit)/NULLIF(SUM(amount),0)*100, 2)
            FROM vw_sales_detail WHERE order_date = CURRENT_DATE - 1
        """,
    },
    {
        "title": "今日销售额 TOP5 客户",
        "sql": """
            SELECT customer_name AS 客户, COUNT(DISTINCT order_no) AS 订单数,
                   ROUND(SUM(amount)::numeric, 2) AS 销售额
            FROM vw_sales_detail WHERE order_date = CURRENT_DATE
            GROUP BY customer_name ORDER BY 销售额 DESC LIMIT 5
        """,
    },
    {
        "title": "今日销售额 TOP5 商品",
        "sql": """
            SELECT product_name AS 商品, brand AS 品牌, SUM(quantity) AS 销量,
                   ROUND(SUM(amount)::numeric, 2) AS 销售额
            FROM vw_sales_detail WHERE order_date = CURRENT_DATE
            GROUP BY product_name, brand ORDER BY 销售额 DESC LIMIT 5
        """,
    },
    {
        "title": "采购概况（环比）",
        "sql": """
            SELECT
                '今日' AS 日期,
                COUNT(DISTINCT po_no) AS 采购单数,
                ROUND(SUM(amount)::numeric, 2) AS 采购金额,
                SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END) AS 已完成,
                SUM(CASE WHEN status='partial' THEN 1 ELSE 0 END) AS 部分到货,
                SUM(CASE WHEN status IN ('pending_review','pending','paid') THEN 1 ELSE 0 END) AS 进行中
            FROM vw_purchase_detail WHERE purchase_date = CURRENT_DATE
            UNION ALL
            SELECT
                '昨日',
                COUNT(DISTINCT po_no),
                ROUND(SUM(amount)::numeric, 2),
                SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END),
                SUM(CASE WHEN status='partial' THEN 1 ELSE 0 END),
                SUM(CASE WHEN status IN ('pending_review','pending','paid') THEN 1 ELSE 0 END)
            FROM vw_purchase_detail WHERE purchase_date = CURRENT_DATE - 1
        """,
    },
    {
        "title": "库存周转 TOP10（最慢）",
        "sql": """
            SELECT product_name AS 商品, brand AS 品牌, current_stock AS 库存,
                   sold_30d AS 近30天销量, ROUND(turnover_rate::numeric, 2) AS 周转率
            FROM vw_inventory_turnover WHERE current_stock > 0
            ORDER BY turnover_rate ASC LIMIT 10
        """,
    },
    {
        "title": "应收账款概况",
        "sql": """
            SELECT COUNT(*) AS 笔数,
                   ROUND(SUM(unreceived_amount)::numeric, 2) AS 未收总额,
                   ROUND(SUM(CASE WHEN age_days > 30 THEN unreceived_amount ELSE 0 END)::numeric, 2) AS 逾期金额
            FROM vw_receivables WHERE status != 'completed'
        """,
    },
    {
        "title": "应付账款概况",
        "sql": """
            SELECT COUNT(*) AS 笔数,
                   ROUND(SUM(unpaid_amount)::numeric, 2) AS 未付总额,
                   ROUND(SUM(CASE WHEN age_days > 30 THEN unpaid_amount ELSE 0 END)::numeric, 2) AS 逾期金额
            FROM vw_payables WHERE status != 'completed'
        """,
    },
    {
        "title": "账龄超30天客户欠款",
        "sql": """
            SELECT customer_name AS 客户, bill_no AS 单号, bill_date AS 账单日期,
                   ROUND(unreceived_amount::numeric, 2) AS 未收金额, age_days AS 账龄天数
            FROM vw_receivables WHERE status != 'completed' AND age_days > 30
            ORDER BY age_days DESC
        """,
    },
]


async def _load_config() -> dict | None:
    """从 system_settings 加载日报配置"""
    settings = await SystemSetting.filter(key__in=CONFIG_KEYS)
    cfg = {s.key: s.value for s in settings}
    if cfg.get("daily_report.enabled") != "true":
        return None
    if not cfg.get("daily_report.recipients") or not cfg.get("daily_report.smtp_host"):
        return None
    return cfg


def _serialize_value(val):
    """将数据库值转为 JSON 友好格式"""
    from decimal import Decimal
    from datetime import date as _date, datetime as _dt
    if isinstance(val, Decimal):
        return float(val)
    if isinstance(val, (_date, _dt)):
        return val.isoformat()
    return val


async def _execute_queries(db_dsn: str) -> list[dict]:
    """执行所有固定 SQL，返回 tables 数组"""
    from app.services.ai_chat_service import get_ai_pool
    tables = []
    pool = await get_ai_pool(db_dsn)
    for q in REPORT_QUERIES:
        try:
            async with pool.acquire() as conn:
                async with conn.transaction(readonly=True):
                    rows = await conn.fetch(q["sql"])
            if rows:
                columns = list(rows[0].keys())
                row_data = [[_serialize_value(r[c]) for c in columns] for r in rows]
                tables.append({"title": q["title"], "columns": columns, "rows": row_data, "row_count": len(rows)})
            else:
                tables.append({"title": q["title"], "columns": [], "rows": [], "row_count": 0})
        except Exception as e:
            logger.warning(f"日报查询失败 [{q['title']}]: {e}")
            tables.append({"title": q["title"], "columns": [], "rows": [], "row_count": 0})
    return tables


async def _generate_analysis(tables: list[dict]) -> str | None:
    """调用 V3 生成分析摘要（best-effort）"""
    try:
        from app.ai.prompt_builder import build_analysis_prompt
        from app.ai.deepseek_client import call_deepseek

        api_key_encrypted = await _get_setting("ai.deepseek.api_key")
        api_key = decrypt_value(api_key_encrypted)
        if not api_key:
            return None

        base_url = await _get_setting("ai.deepseek.base_url") or "https://api.deepseek.com"
        model = await _get_setting("ai.deepseek.model_analysis") or "deepseek-chat"

        analysis_prompt = build_analysis_prompt(None)
        today = datetime.now().strftime("%Y-%m-%d")
        data_summary = f"用户问题: 生成{today}的业务日报\n\n查询了 {len(tables)} 组数据:\n"
        for t in tables:
            data_summary += f"\n### {t['title']} ({t['row_count']} 行)\n"
            if t["columns"]:
                data_summary += f"列: {t['columns']}\n"
                for row in t["rows"][:10]:
                    data_summary += str(row) + "\n"

        result = await call_deepseek(
            messages=[
                {"role": "system", "content": analysis_prompt},
                {"role": "user", "content": data_summary},
            ],
            api_key=api_key,
            base_url=base_url,
            model=model,
            temperature=0.7,
            max_tokens=2048,
            timeout=60.0,
        )
        if result:
            return result.get("analysis") or result.get("message") or str(result)
    except Exception as e:
        logger.warning(f"日报 AI 分析失败（降级为纯数据）: {e}")
    return None


async def _get_setting(key: str) -> str | None:
    s = await SystemSetting.filter(key=key).first()
    return s.value if s else None


def _build_html(tables: list[dict], analysis: str | None, report_date: str) -> str:
    """拼装 HTML 邮件正文"""
    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; color: #1a1a1a; max-width: 800px; margin: 0 auto; padding: 20px; }}
h1 {{ font-size: 20px; border-bottom: 2px solid #3b82f6; padding-bottom: 8px; }}
h2 {{ font-size: 15px; color: #374151; margin-top: 24px; margin-bottom: 8px; }}
table {{ border-collapse: collapse; width: 100%; font-size: 13px; }}
th {{ background: #f3f4f6; padding: 8px 12px; text-align: left; font-weight: 600; border: 1px solid #e5e7eb; }}
td {{ padding: 6px 12px; border: 1px solid #e5e7eb; }}
tr:nth-child(even) {{ background: #f9fafb; }}
.analysis {{ background: #f0f9ff; border-left: 3px solid #3b82f6; padding: 12px 16px; margin: 16px 0; font-size: 14px; line-height: 1.6; }}
.empty {{ color: #9ca3af; font-style: italic; font-size: 13px; }}
.footer {{ margin-top: 32px; padding-top: 12px; border-top: 1px solid #e5e7eb; font-size: 12px; color: #9ca3af; }}
</style></head><body>
<h1>业务日报 — {report_date}（截至发送时刻）</h1>
"""
    if analysis:
        html += f'<div class="analysis">{analysis}</div>\n'

    for t in tables:
        html += f"<h2>{t['title']}</h2>\n"
        if not t["columns"]:
            html += '<p class="empty">暂无数据</p>\n'
            continue
        html += "<table><thead><tr>"
        for col in t["columns"]:
            html += f"<th>{col}</th>"
        html += "</tr></thead><tbody>\n"
        for row in t["rows"][:20]:
            html += "<tr>"
            for cell in row:
                val = cell if cell is not None else "-"
                if isinstance(val, float):
                    val = f"{val:,.2f}"
                html += f"<td>{val}</td>"
            html += "</tr>\n"
        if t["row_count"] > 20:
            html += f'<tr><td colspan="{len(t["columns"])}" style="text-align:center;color:#9ca3af">... 共 {t["row_count"]} 行，完整数据见 Excel 附件</td></tr>\n'
        html += "</tbody></table>\n"

    html += f'<div class="footer">此邮件由 ERP 系统自动发送 · {report_date}</div></body></html>'
    return html


def _build_excel(tables: list[dict]) -> bytes:
    """生成 Excel 附件（多 sheet）"""
    import openpyxl
    wb = openpyxl.Workbook()
    for ti, t in enumerate(tables):
        if ti == 0:
            ws = wb.active
        else:
            ws = wb.create_sheet()
        ws.title = t["title"][:31]
        for col, name in enumerate(t.get("columns", []), 1):
            ws.cell(row=1, column=col, value=name)
        for row_idx, row in enumerate(t.get("rows", []), 2):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def generate_and_send_report(db_dsn: str, force: bool = False):
    """主入口：生成日报并发送邮件"""
    import json
    from app.services.email_service import send_email
    from app.services.operation_log_service import log_operation

    cfg = await _load_config()
    if not cfg:
        return

    today = datetime.now().strftime("%Y-%m-%d")

    # 防重复（force=True 时跳过）
    if not force and cfg.get("daily_report.last_sent_date") == today:
        return

    # 检查 SMTP 密码
    smtp_password = decrypt_value(cfg.get("daily_report.smtp_password", ""))
    if not smtp_password:
        logger.warning("日报发送跳过: SMTP 密码未配置")
        return

    logger.info("开始生成日报...")

    # 1. 执行查询
    tables = await _execute_queries(db_dsn)

    # 2. AI 分析（best-effort）
    analysis = await _generate_analysis(tables)

    # 3. 生成 HTML + Excel
    html = _build_html(tables, analysis, today)
    excel_bytes = _build_excel(tables)

    # 4. 发送邮件
    recipients = json.loads(cfg.get("daily_report.recipients", "[]"))
    from_name = cfg.get("daily_report.from_name", "ERP系统")
    from_email = cfg.get("daily_report.from_email", "")
    subject = f"{from_name} 业务日报 — {today}"

    success_count = 0
    for recipient in recipients:
        try:
            await send_email(
                smtp_host=cfg["daily_report.smtp_host"],
                smtp_port=int(cfg.get("daily_report.smtp_port", "465")),
                smtp_user=cfg.get("daily_report.smtp_user", ""),
                smtp_password=smtp_password,
                from_email=from_email,
                from_name=from_name,
                to_email=recipient.strip(),
                subject=subject,
                html_body=html,
                attachments=[(f"业务日报_{today}.xlsx", excel_bytes)],
            )
            success_count += 1
        except Exception as e:
            logger.error(f"日报发送失败 [{recipient}]: {e}")

    # 5. 记录发送状态
    setting = await SystemSetting.filter(key="daily_report.last_sent_date").first()
    if setting:
        setting.value = today
        await setting.save()
    else:
        await SystemSetting.create(key="daily_report.last_sent_date", value=today)

    logger.info(f"日报发送完成: {success_count}/{len(recipients)} 成功")

    try:
        await log_operation(None, "DAILY_REPORT_SEND", "SYSTEM", None,
            f"日报发送: {success_count}/{len(recipients)} 成功, AI分析: {'有' if analysis else '无'}")
    except Exception:
        pass


async def daily_report_loop(db_dsn: str):
    """后台定时任务循环（类似 auto_backup_loop）"""
    logger.info("日报邮件循环已启动")
    while True:
        await asyncio.sleep(60)
        try:
            cfg = await _load_config()
            if not cfg:
                continue
            send_time = cfg.get("daily_report.send_time", "21:00")
            now = datetime.now()
            hour, minute = map(int, send_time.split(":"))
            if now.hour > hour or (now.hour == hour and now.minute >= minute):
                today = now.strftime("%Y-%m-%d")
                if cfg.get("daily_report.last_sent_date") != today:
                    await generate_and_send_report(db_dsn)
        except Exception as e:
            logger.error(f"日报循环异常: {e}")
