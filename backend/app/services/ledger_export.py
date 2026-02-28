"""账簿 Excel 导出"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


_HEADER_FONT = Font(bold=True, size=10)
_TITLE_FONT = Font(bold=True, size=13)
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT = Alignment(horizontal="right")
_HEADER_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _set_header_row(ws, row_num, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER


def export_general_ledger_excel(data: dict) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "总分类账"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    title_cell = ws.cell(1, 1, value=f"总分类账 — {data['account_code']} {data['account_name']}")
    title_cell.font = _TITLE_FONT
    title_cell.alignment = _CENTER
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    ws.cell(2, 1, value=f"期间: {data['start_period']} 至 {data['end_period']}").alignment = _CENTER
    headers = ["日期", "凭证号", "摘要", "借方", "贷方", "方向", "余额"]
    _set_header_row(ws, 4, headers)
    row = 5
    ws.cell(row, 3, value="期初余额")
    ws.cell(row, 6, value=data["opening_direction"])
    ws.cell(row, 7, value=data["opening_balance"]).alignment = _RIGHT
    row += 1
    for e in data["entries"]:
        ws.cell(row, 1, value=e["date"])
        ws.cell(row, 2, value=e["voucher_no"])
        ws.cell(row, 3, value=e["summary"])
        ws.cell(row, 4, value=e["debit"]).alignment = _RIGHT
        ws.cell(row, 5, value=e["credit"]).alignment = _RIGHT
        ws.cell(row, 6, value=e["direction"])
        ws.cell(row, 7, value=e["balance"]).alignment = _RIGHT
        row += 1
    ws.cell(row, 3, value="本期合计").font = _HEADER_FONT
    ws.cell(row, 4, value=data["period_debit_total"]).alignment = _RIGHT
    ws.cell(row, 5, value=data["period_credit_total"]).alignment = _RIGHT
    row += 1
    ws.cell(row, 3, value="期末余额").font = _HEADER_FONT
    ws.cell(row, 6, value=data["closing_direction"])
    ws.cell(row, 7, value=data["closing_balance"]).alignment = _RIGHT
    widths = [12, 26, 30, 15, 15, 6, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_detail_ledger_excel(data: dict) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "明细分类账"
    has_aux = data.get("aux_customer") or data.get("aux_supplier")
    extra_headers = []
    if data.get("aux_customer"):
        extra_headers.append("客户")
    if data.get("aux_supplier"):
        extra_headers.append("供应商")
    col_count = 7 + len(extra_headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws.cell(1, 1, value=f"明细分类账 — {data['account_code']} {data['account_name']}").font = _TITLE_FONT
    ws.cell(1, 1).alignment = _CENTER
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    ws.cell(2, 1, value=f"期间: {data['start_period']} 至 {data['end_period']}").alignment = _CENTER
    headers = ["日期", "凭证号", "摘要"] + extra_headers + ["借方", "贷方", "方向", "余额"]
    _set_header_row(ws, 4, headers)
    row = 5
    ws.cell(row, 3, value="期初余额")
    ws.cell(row, col_count - 1, value=data["opening_direction"])
    ws.cell(row, col_count, value=data["opening_balance"]).alignment = _RIGHT
    row += 1
    for e in data["entries"]:
        col = 1
        ws.cell(row, col, value=e["date"]); col += 1
        ws.cell(row, col, value=e["voucher_no"]); col += 1
        ws.cell(row, col, value=e["summary"]); col += 1
        if data.get("aux_customer"):
            ws.cell(row, col, value=e.get("customer_name") or ""); col += 1
        if data.get("aux_supplier"):
            ws.cell(row, col, value=e.get("supplier_name") or ""); col += 1
        ws.cell(row, col, value=e["debit"]).alignment = _RIGHT; col += 1
        ws.cell(row, col, value=e["credit"]).alignment = _RIGHT; col += 1
        ws.cell(row, col, value=e["direction"]); col += 1
        ws.cell(row, col, value=e["balance"]).alignment = _RIGHT
        row += 1
    offset = len(extra_headers)
    ws.cell(row, 3, value="本期合计").font = _HEADER_FONT
    ws.cell(row, 4 + offset, value=data["period_debit_total"]).alignment = _RIGHT
    ws.cell(row, 5 + offset, value=data["period_credit_total"]).alignment = _RIGHT
    row += 1
    ws.cell(row, 3, value="期末余额").font = _HEADER_FONT
    ws.cell(row, 6 + offset, value=data["closing_direction"])
    ws.cell(row, 7 + offset, value=data["closing_balance"]).alignment = _RIGHT
    widths = [12, 26, 30] + [14] * len(extra_headers) + [15, 15, 6, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_trial_balance_excel(data: dict) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "科目余额表"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(1, 1, value=f"科目余额表 — {data['period_name']}").font = _TITLE_FONT
    ws.cell(1, 1).alignment = _CENTER
    ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
    ws.cell(3, 1, value="科目编码").font = _HEADER_FONT
    ws.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)
    ws.cell(3, 2, value="科目名称").font = _HEADER_FONT
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=4)
    ws.cell(3, 3, value="期初余额").font = _HEADER_FONT
    ws.cell(3, 3).alignment = _CENTER
    ws.cell(4, 3, value="借方").font = _HEADER_FONT
    ws.cell(4, 4, value="贷方").font = _HEADER_FONT
    ws.merge_cells(start_row=3, start_column=5, end_row=3, end_column=6)
    ws.cell(3, 5, value="本期发生额").font = _HEADER_FONT
    ws.cell(3, 5).alignment = _CENTER
    ws.cell(4, 5, value="借方").font = _HEADER_FONT
    ws.cell(4, 6, value="贷方").font = _HEADER_FONT
    ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=8)
    ws.cell(3, 7, value="期末余额").font = _HEADER_FONT
    ws.cell(3, 7).alignment = _CENTER
    ws.cell(4, 7, value="借方").font = _HEADER_FONT
    ws.cell(4, 8, value="贷方").font = _HEADER_FONT
    for c in range(1, 9):
        ws.cell(3, c).fill = _HEADER_FILL
        ws.cell(4, c).fill = _HEADER_FILL
    row = 5
    for a in data["accounts"]:
        indent = "  " * (a["level"] - 1)
        ws.cell(row, 1, value=a["code"])
        ws.cell(row, 2, value=f"{indent}{a['name']}")
        ws.cell(row, 3, value=a["opening_debit"]).alignment = _RIGHT
        ws.cell(row, 4, value=a["opening_credit"]).alignment = _RIGHT
        ws.cell(row, 5, value=a["period_debit"]).alignment = _RIGHT
        ws.cell(row, 6, value=a["period_credit"]).alignment = _RIGHT
        ws.cell(row, 7, value=a["closing_debit"]).alignment = _RIGHT
        ws.cell(row, 8, value=a["closing_credit"]).alignment = _RIGHT
        if not a["is_leaf"]:
            for c in range(1, 9):
                ws.cell(row, c).font = _HEADER_FONT
        row += 1
    t = data["totals"]
    ws.cell(row, 2, value="合  计").font = Font(bold=True, size=11)
    ws.cell(row, 3, value=t["opening_debit"]).alignment = _RIGHT
    ws.cell(row, 4, value=t["opening_credit"]).alignment = _RIGHT
    ws.cell(row, 5, value=t["period_debit"]).alignment = _RIGHT
    ws.cell(row, 6, value=t["period_credit"]).alignment = _RIGHT
    ws.cell(row, 7, value=t["closing_debit"]).alignment = _RIGHT
    ws.cell(row, 8, value=t["closing_credit"]).alignment = _RIGHT
    for c in range(1, 9):
        ws.cell(row, c).font = _HEADER_FONT
    row += 1
    balanced_text = "试算平衡" if data["is_balanced"] else "试算不平衡！"
    ws.cell(row, 2, value=balanced_text).font = Font(
        bold=True, color="008000" if data["is_balanced"] else "FF0000"
    )
    widths = [12, 24, 14, 14, 14, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
